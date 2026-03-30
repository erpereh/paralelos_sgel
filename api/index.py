from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pandas as pd
import numpy as np
from io import BytesIO
import base64
import json
import os
import re
import unicodedata
import datetime as dt
import difflib
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List
import google.generativeai as genai

app = FastAPI()
_POBLACIONES_CACHE = None
_GEOGRAPHY_AI_CACHE = {}
_GEMINI_MODEL_CACHE = None

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def clean_dni(value) -> str:
    """Normalize DNI/ID: strip whitespace, uppercase, remove leading zeros.
    Handles pandas float conversion (10008.0 -> '10008').
    """
    if pd.isna(value):
        return ""
    # If it's a float that is really an integer (e.g. 10008.0), convert to int first
    if isinstance(value, float) and value == int(value):
        value = int(value)
    s = str(value).strip().upper()
    # Remove dashes and spaces (but NOT dots yet — they were handled above)
    s = s.replace("-", "").replace(" ", "")
    # If there's a trailing ".0" still (from string-typed "10008.0"), remove it
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    # If the entire string is numeric, strip leading zeros
    if s.isdigit():
        s = s.lstrip("0") or "0"
    # If the string ends with a letter but the rest is numeric (like a Spanish DNI),
    # strip leading zeros from the numeric part.
    elif len(s) > 1 and s[:-1].isdigit():
        numeric_part = s[:-1].lstrip("0") or "0"
        s = numeric_part + s[-1]
    return s


def safe_float(val):
    if pd.isna(val):
        return 0.0
    try:
        if isinstance(val, str):
            s = val.strip().replace('.', '').replace(',', '.')
            if s == "":
                return 0.0
            return float(s)
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _normalize_column(col_name: str) -> str:
    return str(col_name).strip().upper().replace('_', '').replace('.', '').replace(' ', '')


def _column_matches(col_name: str, possible_names) -> bool:
    norm = _normalize_column(col_name)
    for p in possible_names:
        if _normalize_column(p) in norm or norm in _normalize_column(p):
            return True
    return False


def read_excel_guess_header(bytes_data, candidates, required_columns):
    """Try reading same file with several skiprows until required columns are found."""
    last_exception = None
    for skip in candidates:
        try:
            df = pd.read_excel(BytesIO(bytes_data), skiprows=skip, engine="openpyxl")
            df.columns = df.columns.astype(str).str.strip().str.replace('\n', ' ').str.replace('\r', '')

            ok = True
            for req in required_columns:
                if not any(_column_matches(col, req) for col in df.columns):
                    ok = False
                    break
            if ok:
                return df

        except Exception as exc:
            last_exception = exc
            continue

    if last_exception:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel con los offsets ({candidates}): {last_exception}")
    raise HTTPException(status_code=400, detail="No se encontró la cabecera esperada en el Excel con los offsets probados.")


def find_column(df, possible_names, required=False, exact=False):
    """
    Find a column in df based on a list of possible names.
    If exact=True, it checks for case-insensitive exact match.
    If exact=False, it checks if any of the possible names are a substring of the column name.
    """
    cols_upper = {c: c.upper() for c in df.columns}
    for p in possible_names:
        p_up = p.upper()
        for orig_col, up_col in cols_upper.items():
            if exact:
                if p_up == up_col:
                    return orig_col
            else:
                if p_up in up_col:
                    return orig_col
    
    if required:
        available = list(df.columns)
        raise HTTPException(
            status_code=400, 
            detail=f"No se encontró una columna válida para {possible_names[0]}. Las columnas disponibles son: {available}"
        )
    return None

# ──────────────────────────────────────────────────────────────
# ENDPOINT 1: /api/process — Upload + clean + drop_duplicates + merge
# ──────────────────────────────────────────────────────────────
@app.post("/api/process")
async def process_payroll(
    file_xrp: UploadFile = File(...),
    file_meta4: UploadFile = File(...),
    match_by: str = Form("dni"),  # accepted: 'dni' or 'id'
):
    try:
        # ==========================================
        # 1. READ AND CLEAN XRP
        # ==========================================
        xrp_bytes = await file_xrp.read()
        df_xrp = read_excel_guess_header(
            xrp_bytes,
            [4, 3, 2, 1, 0],
            required_columns=[
                ["Trabajador", "DNI trabajador", "DNI"],
                ["Nombre trabajador", "Nombre"],
                ["Total Devengado", "Devengos"],
            ],
        )

        # Buscar columnas XRP
        XRP_COL_ID = find_column(df_xrp, ["Trabajador", "DNI trabajador", "DNI"], required=True)
        XRP_COL_NOMBRE = find_column(df_xrp, ["Nombre trabajador", "Nombre"], required=True)
        XRP_COL_DEVENGOS = find_column(df_xrp, ["Total Devengado", "Devengos"], required=True)
        XRP_COL_DEDUCCIONES = find_column(df_xrp, ["Deducciones", "Retenciones", "Retenido"], required=False)
        XRP_COL_LIQUIDO = find_column(df_xrp, ["Liquido", "Neto", "Percibir"], required=False)
        XRP_COL_CONVENIO = find_column(df_xrp, ["Convenio"], required=False)

        xrp_data = pd.DataFrame()
        xrp_data["dni_clean"] = df_xrp[XRP_COL_ID].apply(clean_dni)
        xrp_data["id_xrp"] = df_xrp[XRP_COL_ID].apply(clean_dni)
        xrp_data["nombre_xrp"] = df_xrp[XRP_COL_NOMBRE].astype(str).str.strip()
        xrp_data["devengos_xrp"] = df_xrp[XRP_COL_DEVENGOS].apply(safe_float)
        
        if XRP_COL_DEDUCCIONES:
            xrp_data["deducciones_xrp"] = df_xrp[XRP_COL_DEDUCCIONES].apply(safe_float)
        else:
            xrp_data["deducciones_xrp"] = 0.0
            
        if XRP_COL_LIQUIDO:
            xrp_data["liquido_xrp"] = df_xrp[XRP_COL_LIQUIDO].apply(safe_float)
        else:
            xrp_data["liquido_xrp"] = 0.0
            
        if XRP_COL_CONVENIO:
            xrp_data["convenio_xrp"] = df_xrp[XRP_COL_CONVENIO].apply(clean_dni)
        else:
            xrp_data["convenio_xrp"] = ""

        # Eliminar vacíos y duplicados estrictamente (XRP)
        xrp_data = xrp_data[xrp_data["dni_clean"] != ""]
        xrp_data = xrp_data.drop_duplicates(subset=['dni_clean'], keep='first')


        # ==========================================
        # 2. READ AND CLEAN META4
        # ==========================================
        meta4_bytes = await file_meta4.read()
        df_meta4 = read_excel_guess_header(
            meta4_bytes,
            [3, 2, 1, 0],
            required_columns=[
                ["Empleado", "DNI", "Trabajador"],
                ["Nombre"],
                ["Total_Devengos", "Bruto", "Devengos"],
                ["Total.Retenido", "Deducciones", "Retenciones"],
                ["Liquido", "Neto", "Percibir"],
            ],
        )

        # Buscar columnas Meta4
        # Preferimos la columna Trabajador para el ID de empleado cuando está presente.
        META4_COL_ID = find_column(df_meta4, ["Trabajador", "Empleado", "DNI"], required=True)
        META4_COL_NOMBRE = find_column(df_meta4, ["Nombre"], required=True)
        META4_COL_EMPRESA = find_column(df_meta4, ["Centro_de_Trabajo", "Empresa", "Centro"], required=False)
        META4_COL_DEVENGOS = find_column(df_meta4, ["Total_Devengos", "Bruto", "Devengos"], required=True)
        META4_COL_DEDUCCIONES = find_column(df_meta4, ["Total.Retenido", "Deducciones", "Retenciones"], required=True)
        META4_COL_LIQUIDO = find_column(df_meta4, ["Liquido", "Neto", "Percibir"], required=True)
        META4_COL_CONVENIO = find_column(df_meta4, ["id.Convenio", "Convenio"], required=False)

        meta4_data = pd.DataFrame()
        meta4_data["dni_clean"] = df_meta4[META4_COL_ID].apply(clean_dni)
        meta4_data["id_meta4"] = df_meta4[META4_COL_ID].apply(clean_dni)

        # Si el ID real está en otro campo (p.ej. DNI) y además existe Trabajador, preferimos Trabajador como id_empleado.
        if META4_COL_ID != "Trabajador":
            col_trab = find_column(df_meta4, ["Trabajador"], required=False)
            if col_trab is not None:
                meta4_data["id_meta4"] = df_meta4[col_trab].apply(clean_dni)

        # Nombre en Meta4 (intentando juntar apellidos si existen, o usar nombre directo)
        col_ap1 = find_column(df_meta4, ["Apellido_1", "Primer Apellido"])
        col_ap2 = find_column(df_meta4, ["Apellido_2", "Segundo Apellido"])
        
        if col_ap1:
            n_part = df_meta4[col_ap1].fillna("").astype(str).str.strip() + " "
            if col_ap2:
                n_part += df_meta4[col_ap2].fillna("").astype(str).str.strip() + ", "
            else:
                n_part += ", "
            n_part += df_meta4[META4_COL_NOMBRE].fillna("").astype(str).str.strip()
            
            meta4_data["nombre_meta4"] = n_part.str.strip().str.replace(r"^,\s*", "", regex=True)
        else:
            meta4_data["nombre_meta4"] = df_meta4[META4_COL_NOMBRE].astype(str).str.strip()

        if META4_COL_EMPRESA:
            meta4_data["empresa"] = df_meta4[META4_COL_EMPRESA].fillna("").astype(str).str.strip()
        else:
            meta4_data["empresa"] = ""
            
        meta4_data["devengos_meta4"] = df_meta4[META4_COL_DEVENGOS].apply(safe_float)
        meta4_data["deducciones_meta4"] = df_meta4[META4_COL_DEDUCCIONES].apply(safe_float)
        meta4_data["liquido_meta4"] = df_meta4[META4_COL_LIQUIDO].apply(safe_float)

        if META4_COL_CONVENIO:
            meta4_data["convenio_meta4"] = df_meta4[META4_COL_CONVENIO].apply(clean_dni)
        else:
            meta4_data["convenio_meta4"] = ""

        # Eliminar vacíos y duplicados estrictamente (Meta4)
        meta4_data = meta4_data[meta4_data["dni_clean"] != ""]
        meta4_data = meta4_data.drop_duplicates(subset=['dni_clean'], keep='first')

        # Build merge key (dni o id)
        if match_by == "id":
            xrp_data["match_key"] = xrp_data["id_xrp"].apply(clean_dni)
            meta4_data["match_key"] = meta4_data["id_meta4"].apply(clean_dni)
        else:
            xrp_data["match_key"] = xrp_data["dni_clean"]
            meta4_data["match_key"] = meta4_data["dni_clean"]

        # Avoid duplicates on selected key (when id key is repeated)
        xrp_data = xrp_data.drop_duplicates(subset=["match_key"], keep="first")
        meta4_data = meta4_data.drop_duplicates(subset=["match_key"], keep="first")

        # ==========================================
        # 3. MERGE (OUTER)
        # =========================================
        xrp_ids = set(xrp_data["match_key"].astype(str).str.strip())
        meta4_ids = set(meta4_data["match_key"].astype(str).str.strip())
        common_ids = xrp_ids & meta4_ids

        df_merged = pd.merge(meta4_data, xrp_data, on="match_key", how="outer", indicator=True)

        result_rows = []
        for _, row in df_merged.iterrows():
            merge_status = row["_merge"] 
            
            # Preferir nombre/ID de Meta4 si existe, sino usar el de XRP
            nombre = str(row.get("nombre_meta4", "") if pd.notna(row.get("nombre_meta4")) else row.get("nombre_xrp", "")).strip()
            id_emp = str(row.get("id_meta4", "") if pd.notna(row.get("id_meta4")) else row.get("id_xrp", "")).strip()
            if not id_emp:
                id_emp = str(row["dni_clean"])
            
            empresa = str(row.get("empresa", "") if pd.notna(row.get("empresa")) else "").strip()

            # Forzar NaN a 0.0 mediante safe_float
            dev_xrp = safe_float(row.get("devengos_xrp"))
            ded_xrp = safe_float(row.get("deducciones_xrp"))
            liq_xrp = safe_float(row.get("liquido_xrp"))
            
            dev_m4 = safe_float(row.get("devengos_meta4"))
            ded_m4 = safe_float(row.get("deducciones_meta4"))
            liq_m4 = safe_float(row.get("liquido_meta4"))

            diferencia = round(liq_m4 - liq_xrp, 2)
            
            conv_xrp = str(row.get("convenio_xrp", "") if pd.notna(row.get("convenio_xrp")) else "").strip()
            conv_m4 = str(row.get("convenio_meta4", "") if pd.notna(row.get("convenio_meta4")) else "").strip()
            
            # Match strictly but ignore if both are missing
            if conv_xrp == "" and conv_m4 == "":
                conv_match = ""
            elif conv_xrp == conv_m4:
                conv_match = "COINCIDE"
            else:
                conv_match = "NO COINCIDE"

            result_rows.append({
                "nombre": nombre,
                "id_empleado": id_emp,
                "empresa": empresa,
                "devengos_xrp": round(dev_xrp, 2),
                "deducciones_xrp": round(ded_xrp, 2),
                "liquido_xrp": round(liq_xrp, 2),
                "devengos_meta4": round(dev_m4, 2),
                "deducciones_meta4": round(ded_m4, 2),
                "liquido_meta4": round(liq_m4, 2),
                "diferencia": diferencia,
                "convenio_xrp": conv_xrp,
                "convenio_meta4": conv_m4,
                "convenio_match": conv_match,
                "_merge": str(merge_status)
            })

        rows_with_diff = sum(1 for r in result_rows if abs(r["diferencia"]) > 0.01)

        return JSONResponse(content={
            "data": result_rows,
            "total_rows": len(result_rows),
            "rows_with_diff": rows_with_diff,
            "debug": {
                "match_by": match_by,
                "xrp_rows": len(xrp_data),
                "meta4_rows": len(meta4_data),
                "common_keys": len(common_ids),
                "xrp_only_keys": len(xrp_ids - meta4_ids),
                "meta4_only_keys": len(meta4_ids - xrp_ids),
                "sample_common": list(list(common_ids)[:10]),
                "sample_xrp_only": list(list(xrp_ids - meta4_ids)[:10]),
                "sample_meta4_only": list(list(meta4_ids - xrp_ids)[:10]),
            }
        })

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivos: {str(e)}")


# ──────────────────────────────────────────────────────────────
# ENDPOINT 2: /api/generate-excel
# ──────────────────────────────────────────────────────────────
class RowData(BaseModel):
    nombre: str = ""
    id_empleado: str = ""
    empresa: str = ""
    devengos_xrp: float = 0
    deducciones_xrp: float = 0
    liquido_xrp: float = 0
    devengos_meta4: float = 0
    deducciones_meta4: float = 0
    liquido_meta4: float = 0
    diferencia: float = 0
    convenio_xrp: str = ""
    convenio_meta4: str = ""
    convenio_match: str = ""
    _merge: str = "" # ignore when exporting


class ExcelRequest(BaseModel):
    data: List[RowData]


@app.post("/api/generate-excel")
async def generate_excel(req: ExcelRequest):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativa Nóminas"

        headers = [
            "Nombre", "ID Empleado", "Empresa",
            "Devengos XRP", "Deducciones XRP", "LÍQUIDO XRP",
            "Devengos META4", "Deducciones META4", "LÍQUIDO META4",
            "DIFERENCIA", "CONVENIO XRP", "CONVENIO META4"
        ]

        # Estilos visuales
        header_fill = PatternFill(start_color="1C4CB5", end_color="1C4CB5", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        diff_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        conv_diff_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Orange/Yellow
        data_font = Font(name="Calibri", size=10)
        num_fmt = "#,##0.00"

        # Títulos
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # Datos
        for row_idx, record in enumerate(req.data, 2):
            values = [
                record.nombre, record.id_empleado, record.empresa,
                record.devengos_xrp, record.deducciones_xrp, record.liquido_xrp,
                record.devengos_meta4, record.deducciones_meta4, record.liquido_meta4,
                record.diferencia,
                record.convenio_xrp, record.convenio_meta4
            ]
            has_diff = abs(record.diferencia) > 0.01
            conv_mismatch = record.convenio_match == "NO COINCIDE"

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                # Format numbers
                if isinstance(value, (int, float)) and col_idx >= 4:
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal="right")
                
                # Format background if diff in totals
                if has_diff:
                    cell.fill = diff_fill
                
                # Format background specifically for Convenio columns if they mismatch
                if conv_mismatch and col_idx in (11, 12):
                    cell.fill = conv_diff_fill

        # Ajuste dinámico col_width
        for col_idx in range(1, len(headers) + 1):
            max_len = len(headers[col_idx - 1])
            for row_idx in range(2, len(req.data) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 35)

        ws.freeze_panes = "A2"

        # Convertir a Buffer -> Base64
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode("utf-8")

        return JSONResponse(content={"excel_base64": b64})

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")


def _normalize_header(value: str) -> str:
    value = value.strip().lower()
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", value)


def _extract_target_columns(payload) -> List[str]:
    if isinstance(payload, dict):
        if "columnas" in payload:
            cols = payload["columnas"]
        elif "R0" in payload and isinstance(payload["R0"], dict) and "columnas" in payload["R0"]:
            cols = payload["R0"]["columnas"]
        else:
            cols = []
    elif isinstance(payload, list):
        cols = payload
    else:
        cols = []

    out: List[str] = []
    for c in cols:
        if isinstance(c, dict) and "name" in c:
            out.append(str(c["name"]))
        elif isinstance(c, str):
            out.append(c)
    return out


def _load_sgel_template() -> tuple[List[str], dict, dict]:
    base_dir = os.path.dirname(__file__)
    primary_path = os.path.join(base_dir, "templates", "r_formats.json")
    fallback_path = os.path.join(base_dir, "templates", "sgel_r_format.json")

    if os.path.exists(primary_path):
        with open(primary_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        r0 = payload.get("R0", {}) if isinstance(payload, dict) else {}
        columns = r0.get("columnas", []) if isinstance(r0, dict) else []
        target_cols: List[str] = []
        defaults: dict = {}
        meta_map: dict = {}
        for col in columns:
            if not isinstance(col, dict):
                continue
            name = str(col.get("name", "")).strip()
            if not name:
                continue
            target_cols.append(name)
            meta_map[name] = col
            if "default" in col:
                defaults[name] = col.get("default")
        if target_cols:
            return target_cols, defaults, meta_map

    if os.path.exists(fallback_path):
        with open(fallback_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        cols = _extract_target_columns(payload)
        if cols:
            return cols, {}, {}

    raise HTTPException(status_code=500, detail="No se encontrÃ³ un archivo de referencia SGEL vÃ¡lido.")


def _heuristic_mapping(source_headers: List[str], target_headers: List[str]) -> dict:
    source_map = { _normalize_header(h): h for h in source_headers }
    mapping = {}
    for target in target_headers:
        t_norm = _normalize_header(target)
        if t_norm in source_map:
            mapping[target] = source_map[t_norm]
            continue
        # Partial contains matching as fallback
        found = ""
        for s_norm, s_raw in source_map.items():
            if t_norm and (t_norm in s_norm or s_norm in t_norm):
                found = s_raw
                break
        mapping[target] = found
    return mapping


def _detect_header_row(df_raw: pd.DataFrame, max_rows: int = 20) -> int:
    best_row = 0
    best_score = -1
    rows_to_check = min(max_rows, len(df_raw))
    for i in range(rows_to_check):
        row = df_raw.iloc[i].tolist()
        non_empty = [c for c in row if isinstance(c, str) and c.strip()]
        score = len(non_empty)
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _is_date_like(value) -> bool:
    if pd.isna(value):
        return False
    return isinstance(value, (pd.Timestamp, dt.date, dt.datetime))


def _normalize_name_for_checks(value: str) -> str:
    return _normalize_header(value)


def _clean_code(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _pad_code(value, width: int) -> str:
    code = _clean_code(value)
    if not code:
        return ""
    if code.isdigit():
        return code.zfill(width)
    return code


def _split_semicolon_name(value) -> str:
    text = _clean_code(value)
    if ";" not in text:
        return text
    return text.split(";", 1)[1].strip()


def _geography_name_variants(value: str) -> set[str]:
    text = str(value or "").strip()
    if not text:
        return set()

    variants = {text}

    lowered = text.lower()
    lowered = lowered.replace("l'", "").replace("d'", "").replace("de l'", " ")
    variants.add(lowered)

    # Expand frequent municipality abbreviations seen in payroll exports.
    expanded = lowered
    replacements = {
        "s. ": "san ",
        "s ": "san ",
        "seb.": "sebastian",
        "seb ": "sebastian ",
        "sta.": "santa",
        "sta ": "santa ",
        "sto.": "santo",
        "sto ": "santo ",
    }
    for source, target in replacements.items():
        expanded = expanded.replace(source, target)
    expanded = re.sub(r"\s+", " ", expanded).strip()
    variants.add(expanded)

    simplified = lowered
    for token in [" de ", " del ", " de la ", " de las ", " de los ", " la ", " las ", " los ", " el "]:
        simplified = simplified.replace(token, " ")
    simplified = re.sub(r"\s+", " ", simplified).strip()
    variants.add(simplified)

    simplified_expanded = expanded
    for token in [" de ", " del ", " de la ", " de las ", " de los ", " la ", " las ", " los ", " el "]:
        simplified_expanded = simplified_expanded.replace(token, " ")
    simplified_expanded = re.sub(r"\s+", " ", simplified_expanded).strip()
    variants.add(simplified_expanded)

    no_spaces = simplified.replace(" ", "")
    if no_spaces:
        variants.add(no_spaces)

    no_spaces_expanded = simplified_expanded.replace(" ", "")
    if no_spaces_expanded:
        variants.add(no_spaces_expanded)

    return {variant for variant in variants if variant.strip()}


def _load_poblaciones_lookup():
    global _POBLACIONES_CACHE
    if _POBLACIONES_CACHE is not None:
        return _POBLACIONES_CACHE

    base_dir = os.path.dirname(__file__)
    path = os.path.join(base_dir, "..", "public", "POBLACIONES.xlsx")
    path = os.path.abspath(path)

    if not os.path.exists(path):
        _POBLACIONES_CACHE = {
            "localities": {},
            "country_aliases": {"espana": "011", "spain": "011", "es": "011"},
        }
        return _POBLACIONES_CACHE

    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [str(col).strip() if pd.notna(col) else "" for col in df.columns]

    records_by_locality = {}
    province_entries = {}
    localities_by_province = {}
    for _, row in df.iterrows():
        country_code = _pad_code(row.iloc[0] if len(row) > 0 else "", 3)
        province_code = _pad_code(row.iloc[1] if len(row) > 1 else "", 2)
        locality_with_prefix = row.iloc[2] if len(row) > 2 else ""
        population_code = _pad_code(row.iloc[3] if len(row) > 3 else "", 3)
        population_with_prefix = row.iloc[4] if len(row) > 4 else ""
        locality_name = str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else ""
        cp_value = _clean_code(row.iloc[6] if len(row) > 6 else "")
        province_name = _split_semicolon_name(locality_with_prefix)

        candidate_names = {
            locality_name,
            _split_semicolon_name(locality_with_prefix),
            _split_semicolon_name(population_with_prefix),
        }
        entry = {
            "country_code": country_code,
            "province_code": province_code,
            "population_code": population_code,
            "province_name": province_name,
            "locality_name": locality_name,
            "cp": cp_value,
        }
        province_key = _normalize_header(province_name)
        if province_key and province_key not in province_entries:
            province_entries[province_key] = {
                "province_code": province_code,
                "province_name": province_name,
                "country_code": country_code,
            }
        if province_code and locality_name:
            localities_by_province.setdefault(province_code, {})
            localities_by_province[province_code][_normalize_header(locality_name)] = entry
        for name in candidate_names:
            for variant in _geography_name_variants(name):
                norm = _normalize_header(variant)
                if not norm:
                    continue
                records_by_locality.setdefault(norm, []).append(entry)

    provinces_by_name = {}
    countries_by_name = {"espana": "011", "spain": "011", "es": "011"}
    for entries in records_by_locality.values():
        for entry in entries:
            province_norm = _normalize_header(entry["province_name"])
            if province_norm and province_norm not in provinces_by_name:
                provinces_by_name[province_norm] = entry["province_code"]

    _POBLACIONES_CACHE = {
        "localities": records_by_locality,
        "province_aliases": provinces_by_name,
        "country_aliases": countries_by_name,
        "province_entries": province_entries,
        "localities_by_province": localities_by_province,
    }
    return _POBLACIONES_CACHE


def _pick_locality_entry(locality_value, province_hint=""):
    lookup = _load_poblaciones_lookup()
    province_hint_clean = _clean_code(province_hint)

    for variant in _geography_name_variants(locality_value):
        norm_locality = _normalize_header(variant)
        if not norm_locality:
            continue

        matches = lookup["localities"].get(norm_locality, [])
        if not matches:
            continue

        if province_hint_clean:
            for match in matches:
                if match["province_code"] == province_hint_clean:
                    return match
        return matches[0]

    normalized_variants = []
    for variant in _geography_name_variants(locality_value):
        norm = _normalize_header(variant)
        if norm:
            normalized_variants.append(norm)

    if not normalized_variants:
        return None

    if province_hint_clean and province_hint_clean in lookup["localities_by_province"]:
        candidate_keys = list(lookup["localities_by_province"][province_hint_clean].keys())
        candidate_map = lookup["localities_by_province"][province_hint_clean]
    else:
        candidate_keys = list(lookup["localities"].keys())
        candidate_map = {key: entries[0] for key, entries in lookup["localities"].items() if entries}

    for norm_variant in normalized_variants:
        close = difflib.get_close_matches(norm_variant, candidate_keys, n=1, cutoff=0.82)
        if close:
            return candidate_map.get(close[0])
    return None


def _get_gemini_text(prompt: str) -> str:
    global _GEMINI_MODEL_CACHE
    api_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not api_key:
        return ""

    genai.configure(api_key=api_key)
    model_candidates = []
    if _GEMINI_MODEL_CACHE is False:
        return ""
    if isinstance(_GEMINI_MODEL_CACHE, str) and _GEMINI_MODEL_CACHE:
        model_candidates.append(_GEMINI_MODEL_CACHE)
    else:
        env_model = os.getenv("GEMINI_MODEL")
        if env_model:
            model_candidates.append(env_model)
        model_candidates.extend([
            "gemini-2.0-flash",
            "gemini-2.0-flash-lite",
            "gemini-1.5-flash",
            "gemini-1.5-pro",
        ])

    last_error = None
    for model_name in model_candidates:
        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            text = response.text or ""
            if text:
                _GEMINI_MODEL_CACHE = model_name
                return text
        except Exception as e:
            last_error = e
            continue

    if last_error:
        print(f"[gemini] model error: {last_error}")
    _GEMINI_MODEL_CACHE = False
    return ""


def _get_gemini_json(prompt: str):
    text = _get_gemini_text(prompt)
    if not text:
        return None

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None

    try:
        data = json.loads(text[start : end + 1])
    except json.JSONDecodeError:
        return None
    return data if isinstance(data, dict) else None


def _top_candidate_names(raw_value, candidate_names, limit: int = 12) -> List[str]:
    normalized_to_display = {}
    for name in candidate_names:
        norm = _normalize_header(name)
        if norm and norm not in normalized_to_display:
            normalized_to_display[norm] = name

    raw_norm = _normalize_header(str(raw_value or ""))
    if not raw_norm:
        return []

    matches = difflib.get_close_matches(raw_norm, list(normalized_to_display.keys()), n=limit, cutoff=0.45)
    return [normalized_to_display[m] for m in matches]


def _should_try_geography_ai(raw_value) -> bool:
    text = str(raw_value or "").strip()
    if not text:
        return False

    norm = _normalize_header(text)
    if not norm:
        return False
    if norm in {"desconocido", "unknown", "na", "n/a", "null", "none", "nan"}:
        return False
    if text.isdigit():
        return False
    return True


def _resolve_candidate_with_gemini(kind: str, raw_value, candidate_names: List[str], extra_context: str = "") -> str:
    cache_key = f"{kind}:{_normalize_header(str(raw_value or ''))}:{extra_context}"
    if cache_key in _GEOGRAPHY_AI_CACHE:
        return _GEOGRAPHY_AI_CACHE[cache_key]

    if not candidate_names or not _should_try_geography_ai(raw_value):
        _GEOGRAPHY_AI_CACHE[cache_key] = ""
        return ""

    prompt = (
        f"You are matching a Spanish {kind} name to an official candidate list.\n"
        "Return ONLY valid JSON with this shape: {\"match\": \"exact candidate from the list or empty string\"}.\n"
        "Pick the closest equivalent even if spelling, language, accents, or old/new naming differs.\n"
        "If no candidate is credible, return an empty string.\n\n"
        f"Raw value: {raw_value}\n"
        f"Context: {extra_context}\n"
        f"Candidates: {candidate_names}\n"
    )
    data = _get_gemini_json(prompt) or {}
    match = str(data.get("match", "")).strip()
    if match not in candidate_names:
        match = ""
    _GEOGRAPHY_AI_CACHE[cache_key] = match
    return match


def _apply_geography_codes(out: pd.DataFrame):
    lookup = _load_poblaciones_lookup()
    geography_groups = [
        ("DesPais", "DesProvincia", "DesPoblacion"),
    ]

    for country_col, province_col, population_col in geography_groups:
        if population_col not in out.columns:
            continue

        for idx in out.index:
            try:
                country_value = out.at[idx, country_col] if country_col in out.columns else ""
                province_value = out.at[idx, province_col] if province_col in out.columns else ""
                population_value = out.at[idx, population_col]

                entry = _pick_locality_entry(population_value, province_value)
                if entry:
                    if country_col in out.columns:
                        out.at[idx, country_col] = entry["country_code"]
                    if province_col in out.columns:
                        out.at[idx, province_col] = entry["province_code"]
                    out.at[idx, population_col] = entry["population_code"]
                    continue

                norm_country = _normalize_header(str(country_value or ""))
                if country_col in out.columns and norm_country in lookup["country_aliases"]:
                    out.at[idx, country_col] = lookup["country_aliases"][norm_country]

                norm_province = _normalize_header(str(province_value or ""))
                if province_col in out.columns and norm_province in lookup["province_aliases"]:
                    out.at[idx, province_col] = lookup["province_aliases"][norm_province]

                resolved_province_code = _clean_code(out.at[idx, province_col]) if province_col in out.columns else ""
                if province_col in out.columns and not resolved_province_code and str(province_value or "").strip():
                    province_candidates = [entry["province_name"] for entry in lookup["province_entries"].values()]
                    best_province = _resolve_candidate_with_gemini("province", province_value, province_candidates)
                    province_key = _normalize_header(best_province)
                    province_entry = lookup["province_entries"].get(province_key)
                    if province_entry:
                        out.at[idx, province_col] = province_entry["province_code"]
                        resolved_province_code = province_entry["province_code"]
                        if country_col in out.columns and not _clean_code(out.at[idx, country_col]):
                            out.at[idx, country_col] = province_entry["country_code"]

                if not _pick_locality_entry(population_value, resolved_province_code):
                    locality_candidates_map = lookup["localities_by_province"].get(resolved_province_code, {})
                    if not locality_candidates_map:
                        locality_candidates_map = {
                            norm: entries[0]
                            for norm, entries in lookup["localities"].items()
                            if entries
                        }
                    locality_candidates = [entry["locality_name"] for entry in locality_candidates_map.values()]
                    best_locality = _resolve_candidate_with_gemini(
                        "locality",
                        population_value,
                        _top_candidate_names(population_value, locality_candidates) or locality_candidates[:12],
                        extra_context=f"province_code={resolved_province_code or 'unknown'}",
                    )
                    if best_locality:
                        locality_norm = _normalize_header(best_locality)
                        locality_entry = locality_candidates_map.get(locality_norm)
                        if locality_entry:
                            if country_col in out.columns:
                                out.at[idx, country_col] = locality_entry["country_code"]
                            if province_col in out.columns:
                                out.at[idx, province_col] = locality_entry["province_code"]
                            out.at[idx, population_col] = locality_entry["population_code"]
                            continue

                population_as_code = _pad_code(population_value, 3)
                if population_as_code and population_as_code.isdigit() and len(population_as_code) == 3:
                    out.at[idx, population_col] = population_as_code

                if country_col in out.columns:
                    out.at[idx, country_col] = _pad_code(out.at[idx, country_col], 3) or out.at[idx, country_col]
                if province_col in out.columns:
                    out.at[idx, province_col] = _pad_code(out.at[idx, province_col], 2) or out.at[idx, province_col]
            except Exception as row_error:
                print(f"[sgel] geography row error idx={idx} country_col={country_col} province_col={province_col} population_col={population_col}: {row_error}")
                continue

    birth_columns = ["DesPaisNacim", "DesProvinciaNacim", "DesPoblacionNacim"]
    base_columns = ["DesPais", "DesProvincia", "DesPoblacion"]
    if all(col in out.columns for col in base_columns) and all(col in out.columns for col in birth_columns):
        out["DesPaisNacim"] = out["DesPais"]
        out["DesProvinciaNacim"] = out["DesProvincia"]
        out["DesPoblacionNacim"] = out["DesPoblacion"]


def _collect_geography_issues(out: pd.DataFrame):
    geography_specs = [
        ("DesPais", 3),
        ("DesProvincia", 2),
        ("DesPoblacion", 3),
        ("DesPaisNacim", 3),
        ("DesProvinciaNacim", 2),
        ("DesPoblacionNacim", 3),
    ]
    issues = []

    for col, expected_width in geography_specs:
        if col not in out.columns:
            continue

        for idx, value in out[col].items():
            if pd.isna(value):
                continue
            code = _clean_code(value)
            if code.isdigit() and len(code) == expected_width:
                continue
            issues.append({
                "fila": int(idx) + 1,
                "columna": col,
                "valor": value,
                "motivo": f"No se pudo convertir {col} a codigo",
            })

    return issues


def _map_headers_with_gemini(source_headers: List[str], target_headers: List[str]) -> dict:
    api_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="Missing GOOGLE_API_KEY or GEMINI_API_KEY for Gemini.")

    prompt = (
        "You map Excel headers to target headers.\n"
        "Return ONLY valid JSON: {\"TargetColumn\": \"SourceColumn\"}.\n"
        "If no match, return empty string.\n\n"
        f"Source headers: {source_headers}\n"
        f"Target headers: {target_headers}\n"
    )

    text = _get_gemini_text(prompt)
    if not text:
        return {}

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return {}

    try:
        data = json.loads(text[start : end + 1])
    except json.JSONDecodeError:
        return {}

    if not isinstance(data, dict):
        return {}
    return {str(k): (str(v) if v is not None else "") for k, v in data.items()}


@app.post("/api/generate-sgel-r")
async def generate_sgel_r(file: UploadFile = File(...)):
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="Archivo no vÃ¡lido.")

        target_columns, defaults, meta_map = _load_sgel_template()

        raw = await file.read()
        df_raw = pd.read_excel(BytesIO(raw), engine="openpyxl", header=None)
        header_row = _detect_header_row(df_raw)
        headers = df_raw.iloc[header_row].tolist()
        headers = [str(h).strip() if pd.notna(h) else "" for h in headers]
        df = df_raw.iloc[header_row + 1 :].copy()
        df.columns = headers
        df = df.loc[:, [c for c in df.columns if c]]
        df = df.dropna(how="all")
        source_headers = [str(c) for c in df.columns]
        print(f"[sgel] detected header row: {header_row}")
        print(f"[sgel] source headers: {source_headers}")
        print(f"[sgel] source rows: {len(df)}")
        if len(df) > 0:
            print(f"[sgel] first row sample: {df.iloc[0].to_dict()}")

        mapping = _map_headers_with_gemini(source_headers, target_columns)
        if mapping:
            print(f"[sgel] gemini mapping: {mapping}")
        if not mapping:
            mapping = _heuristic_mapping(source_headers, target_columns)
            print(f"[sgel] heuristic mapping: {mapping}")

        normalized_sources = { _normalize_header(c): c for c in df.columns }

        out = pd.DataFrame(index=df.index)
        for target in target_columns:
            if target in defaults:
                out[target] = defaults[target]
                continue
            source = mapping.get(target, "")
            if source in df.columns:
                out[target] = df[source]
            else:
                alt = normalized_sources.get(_normalize_header(source), "")
                if alt in df.columns:
                    out[target] = df[alt]
                elif target in defaults:
                    out[target] = defaults[target]
                else:
                    out[target] = pd.NA

        _apply_geography_codes(out)

        issues = _collect_geography_issues(out)
        for col in out.columns:
            meta = meta_map.get(col, {})
            expected_type = str(meta.get("type", "string")).lower()
            series = out[col]

            if expected_type == "date":
                parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
                invalid = series.notna() & parsed.isna()
                for idx, val in series[invalid].items():
                    issues.append({
                        "fila": int(idx) + 1,
                        "columna": col,
                        "valor": val,
                        "motivo": "Fecha no valida"
                    })
                fmt = meta.get("format", "%d/%m/%Y")
                out[col] = parsed.dt.strftime(fmt)
            elif expected_type == "number":
                num = pd.to_numeric(series, errors="coerce")
                invalid = series.notna() & num.isna()
                for idx, val in series[invalid].items():
                    issues.append({
                        "fila": int(idx) + 1,
                        "columna": col,
                        "valor": val,
                        "motivo": "Numero no valido"
                    })
                out[col] = num
            else:
                name_norm = _normalize_name_for_checks(col)
                if "fecha" not in name_norm:
                    invalid = series.apply(_is_date_like)
                    for idx, val in series[invalid].items():
                        issues.append({
                            "fila": int(idx) + 1,
                            "columna": col,
                            "valor": val,
                            "motivo": "Valor de fecha en columna no fecha"
                        })
                    out.loc[invalid, col] = pd.NA
        if len(out) > 0:
            print(f"[sgel] output first row sample: {out.iloc[0].to_dict()}")

        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            out.to_excel(writer, index=False, sheet_name="Plantilla R")
            if issues:
                pd.DataFrame(issues).to_excel(writer, index=False, sheet_name="Validaciones")
        buf.seek(0)

        headers = {
            "Content-Disposition": "attachment; filename=Plantilla_R_Resultado.xlsx"
        }
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException:
        raise
    except Exception as e:
        print("[sgel] fatal error while generating R")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error generando SGEL R: {str(e)}")
        
